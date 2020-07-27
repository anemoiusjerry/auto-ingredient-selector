from .FormulationFiller import FormulationFiller
from config.configParser import FigMe
from ..dlx3 import DLX
import pandas as pd
import re
import math
import os
import copy
import xlsxwriter
from PySide2.QtCore import QObject, Signal, Slot
from datetime import date
from collections import defaultdict as dd
from openpyxl import load_workbook
from BespokeAutoSystem.WarningRaiser import *

class IngredientSelector(QObject):
    launched = Signal(int)
    stateChanged = Signal(str, str, int)
    cancel = Signal()
    error = Signal(str)
    def __init__(self, orders, ingredients, qnair, catalog):
        # signal setup
        QObject.__init__(self)
        self.warn = WarningRaiser()
        config = FigMe()

        self.config = config
        self.SOF = 7

        # orders columns
        self.customerCol = config.getColname("Orders Spreadsheet", "customer")
        self.orderCol = config.getColname("Orders Spreadsheet", "order number")
        self.emailCol = config.getColname("Orders Spreadsheet", "email")
        self.oitemCol = config.getColname("Orders Spreadsheet", "item")

        # ingredients columns
        self.inameCol = config.getColname("Ingredients Spreadsheet", "name")
        self.typeCol = config.getColname("Ingredients Spreadsheet", "type")
        self.skinProbCol = config.getColname("Ingredients Spreadsheet", "skin problem")
        self.contrainsCol = config.getColname("Ingredients Spreadsheet", "contraindications")
        self.stockCol = config.getColname("Ingredients Spreadsheet", "stock")
        self.comedogenicCol = config.getColname("Ingredients Spreadsheet", "comedogenic")
        self.viscocityCol = config.getColname("Ingredients Spreadsheet", "viscosity")
        self.absorptionCol = config.getColname("Ingredients Spreadsheet", "absorption")
        self.EOnoteCol = config.getColname("Ingredients Spreadsheet", "EO note")

        # questionnaire columns
        self.qnameCol = config.getColname("Customer Questionnaire", "name")
        self.qemailCol = config.getColname("Customer Questionnaire", "email")
        self.ailmentCols = config.getColname("Customer Questionnaire", "skin problem")
        self.pregnancyCol = config.getColname("Customer Questionnaire", "pregnancy")
        self.cusContrainCol = config.getColname("Customer Questionnaire", "contraindications")

        # catalog columns
        self.itemCol = config.getColname("Product Catalog", "item")
        self.productCol = config.getColname("Product Catalog", "products")

        # Constants and values
        self.comeConst = config.getConst("comedogenicRating")
        self.viscConst = config.getConst("viscosity")
        self.absorbConst = config.getConst("absorbency")
        self.lowBound = config.getVal("lowBound")
        self.upBound = config.getVal("upBound")
        self.maxupBound = config.getVal("maxupBound")
        self.tpyeoverlap_low = config.getVal("tpyeoverlap_low")
        self.typeoverlap_up = config.getVal("typeoverlap_up")
        self.maxSols = config.getVal("maxsols")
        self.fitWeight = config.getVal("fitweight")
        self.numIngredWeight = config.getVal("numingredweight")
        self.addedBenefitWeight = config.getVal("addedbenefitweight") # need to finish this part

        # initialising the dataframes
        self.orders = orders
        self.ingredients = ingredients
        self.qnair = qnair
        self.catalog = catalog

        # progress variables
        self.solsFound = 0
        self.stop = False

    def selectIngredients(self):
        allErrorString = ""

        # send the launched signal with the length of the orders
        self.launched.emit(self.orders.shape[0])

        # Creating new file for the orders to be saved into
        savedir = self.config.getDir("Export Directory")
        # parentFolderPath = savedir + "/" + "Sheets"
        # if not os.path.exists(parentFolderPath):
        #     os.makedirs(parentFolderPath)

        returns = []
        for index, order in self.orders.iterrows():
            self.progress = index

            # Find the customer questionnaire using the name or email address
            # !!!! NOTE: A dialog should be added to check that the email corresponds to the correct person
            name = order[self.customerCol]
            email = order[self.emailCol]
            # Change the state of the progress dialog
            text = "Order " + order[self.orderCol] +", "+ name
            self.stateChanged.emit("retrieve", text, self.progress)

            if name in self.qnair.index.tolist():
                qdata = self.qnair.loc[name,:]
            elif email in self.qnair[self.qemailCol].values.tolist():
                # Add a dialog that asks if the customer name is indeed the corect customer linked to the email address
                qdata = self.qnair.loc[self.qnair[self.qemailCol].values.tolist().index(email)]
            else:
                self.warn.displayWarningDialog("Questionnaire Retrieval Error", f"No questionaire found for {name}.\n Make sure the names are the same for the Order and Questionnaire.")
                print("no matching name found for ", name)
                continue

            # Finding the products required to fulfil the order. if they cannot be found, skip to next order
            item = order[self.oitemCol]

            try:
                products = self.catalog.loc[item,self.productCol]
            except:
                # add a check to make sure that all the products are within the known products
                #self.warn.displayWarningDialog("Not Found Error", f"Ordered product named {item} not found in catalog.")
                #raise Exception("Unknown Item")
                allErrorString += f"Ordered product named {item} not found in catalog.\n"
                continue

            # Create a folder for the order
            customer_name = str(order[self.customerCol]).title()

            ordername =  str(date.today().strftime("%d.%m.%Y")) + f" - {customer_name} - Skin Analysis"
            #orderFolderName = parentFolderPath + "/" + ordername
            orderFolderName = savedir + "/" + customer_name
            if not os.path.exists(orderFolderName):
                os.makedirs(orderFolderName)

            for product in products:
                # Get the solutions
                # create a new excel workbook for the order
                wbookname = orderFolderName + "/" + ordername + "-" + str(product) + ".xlsx"
                workbook = xlsxwriter.Workbook(wbookname)
                solutions, rows, cols, unresolved = self.orderParser(product, qdata)
                if self.stop:
                    workbook.close()
                    return None
                # create a new worksheet for each solution
                self.writeToWorkbook(workbook, solutions, rows, cols, unresolved)
                workbook.close()

                for solution in solutions:
                    returns.append({"Ingredients": solution[0],
                                    "CustomerName": name,
                                    "ProductType": product})

        if returns:
            if allErrorString != "":
                self.error.emit(allErrorString)
            return returns
        return None

    def writeToWorkbook(self, workbook, solutions, rows, cols, unresolved):

        # Ailment label format
        _ailDict = {"bold": True,
                    "align": "right",
                    "bg_color": "#E696AE"}
        ailment_format = workbook.add_format(_ailDict)

        # Skin problem and Ailment header format
        _hailDict = {"bold": True,
                    "align": "center",
                    "bg_color": "#DB7093",
                    "bottom": True}
        headail_format = workbook.add_format(_hailDict)

        # Ingredient label format
        _ingDict = {"bold": True,
                    "align": "right",
                    "bg_color": "#C71585",
                    "font_color": "white"}
        ingred_format = workbook.add_format(_ingDict)

        # Nodes format
        _nodDict = {"bold": True,
                    "align": "center",
                    "bg_color": "#D8BFD8"}
        node_format = workbook.add_format(_nodDict)

        i=1
        for solution in solutions:
            #sorting the ingredients to be grouped by their type
            sortDict = dd(list)
            tmplst = []
            for ingrd in solution[0]:
                for r in rows:
                    if r[1] == ingrd:
                        type = cols[r[0][-1][0]][0]
                        break
                sortDict[type].append(ingrd)

            for key, val in sortDict.items():
                tmplst.extend(val)
            solution = list(solution)
            solution[0] = tmplst

            worksheet = workbook.add_worksheet("Solution " + str(i))
            # Write the row headers (skin problems & ingredient types)
            row = 2
            col = 0
            worksheet.write(1,col,"SKIN PROBLEMS",headail_format)
            _strLen = 0
            check = True
            _nrow = 0
            for problem in cols:
                if problem[0] not in ["aqueous base","aqueous high performance","anhydrous high performance","anhydrous base","essential oil"]:
                    _strLen = len(problem[0]) if len(problem[0]) > _strLen else _strLen
                    _problem = problem[0][0].upper() + problem[0][1:] # capitalise first letter
                    worksheet.write(row, col, _problem,ailment_format)
                elif check:
                    _nrow = row
                    check=False
                    row = row + 1
                    worksheet.write(row, col, "INGREDIENT TYPE", headail_format)
                    row = row + 1
                    _problem = problem[0][0].upper() + problem[0][1:] # capitalise first letter
                    worksheet.write(row, col, _problem,ailment_format)
                else:
                    _strLen = len(problem[0]) if len(problem[0]) > _strLen else _strLen
                    _problem = problem[0][0].upper() + problem[0][1:] # capitalise first letter
                    worksheet.write(row, col, _problem,ailment_format)
                row = row+1

            worksheet.set_column(col, col, round(_strLen*1))
            # Write the headings (ingredient names) and populate nodes
            hrow = 1
            hcol = 1
            nrow = 2
            ncol = 1
            for ingredient in solution[0]:
                # write the heading
                _ingredient = ingredient[0].upper() + ingredient[1:] # capitalise first letter
                worksheet.write(hrow, hcol, _ingredient,ingred_format)
                worksheet.set_column(hcol, hcol, round(len(ingredient)*1))
                hcol = hcol+1

                # populate the nodes
                for _row in rows:
                    if _row[1] == ingredient:
                        for node in _row[0]:
                            if node[0] + 2 < _nrow:
                                nrow = node[0] + 2
                            else:
                                nrow = node[0] + 4
                            worksheet.write(nrow, ncol, "X", node_format)
                        break

                for benefits in solution[2]:
                    if benefits[0] == ingredient:
                        worksheet.write(row + 2, ncol, "Xtra Benefits",headail_format)
                        j=0
                        for benefit in benefits[1]:
                            worksheet.write(row+3+j, ncol, benefit)
                            j += 1
                        break
                ncol = ncol + 1
            i = i+1

            # Adding the unresolved columns and additional benefits
            row = row + 2
            worksheet.write(row, 0, "Unresolved Conditions", headail_format)
            if len(unresolved) > 0:
                for j in range(len(unresolved)):
                    worksheet.write(row + j + 1, 0, unresolved[j])
            else:
                worksheet.write(row + 1, 0, 'everything is resolved')

            """row = row + 2 + len(unresolved)
            worksheet.write(row, 0, "Additional Benefits", headail_format)
            if len(solution[2]) > 0:
                for j in range(len(solution[2])):
                    worksheet.write(row + 1 + j, 0, solution[2][j])
            else:
                worksheet.write(row + 1, 0, 'No additional benefits')"""

    def orderParser(self, product, qdata):

        types = self.config.getProduct(product, "types")
        # Retrieving the skin problems from customer info
        ailments = [a for col in self.ailmentCols for a in qdata[col] if a]
        # Finding the customers contraindications
        usercons = qdata[self.pregnancyCol] + qdata[self.cusContrainCol]

        # Retrieve the rows and columns that will make up the dlx matrix
        rows, cols = self.matrixGen(product, ailments, usercons)

        # Convert cols into dlx useable format
        cols = [(cols[i],0,self.lowBound,self.upBound) for i in range(len(cols))]
        last = len(cols) - 1

        # If an ingredient type is part of the product recipe, make sure they are included at least once
        for type in types:
            cols.append((type,0,self.tpyeoverlap_low,self.typeoverlap_up))
            colind = len(cols) - 1
            for row in rows:
                ingredtype = self.ingredients.loc[row[1],self.typeCol]
                if type in ingredtype:
                    row[0].append((colind, None))

        # Check that all cols can be stisfied at least once, remove cols that cant be
        colsCovered = set([node[0] for row in rows for node in row[0]])
        unresolved = []
        for col in range(len(cols)-1, -1, -1):
            if col not in colsCovered:
                if col <= last:
                    last = last - 1
                unresolved.append(cols.pop(col)[0])
                for i in range(len(rows)):
                    rows[i] = list(rows[i])
                    tmp = []
                    for node in rows[i][0]:
                        if node[0] != col:
                            if node[0] > col:
                                tmp.append((node[0]-1, node[1]))
                            else:
                                tmp.append(node)
                    rows[i] = (tmp, rows[i][1])

        # check the amount of ingredient types can fulfil the tpyeoverlap_low constraint
        mintype = dd(int)
        mining = dd(int)
        for row in rows:
            for node in row[0]:
                if node[0] > last:
                    mintype[node[0]] = mintype[node[0]] + 1
                else:
                    mining[node[0]] = mining[node[0]] + 1

        for key, val in mintype.items():
            if val < self.tpyeoverlap_low:
                cols[key] = list(cols[key])
                cols[key][2] = val
                cols[key] = tuple(cols[key])

        for key, val in mining.items():
            if val < self.lowBound:
                cols[key] = list(cols[key])
                cols[key][2] = val
                cols[key] = tuple(cols[key])

        # Run the DLX to find all the solutions
        matrix = DLX(cols, rows)
        solutions = self.solve(matrix)

        # Run the DLX with an increased upper bound until max is reached or enough solutions are found
        while len(solutions) < 100 and self.typeoverlap_up <= len(rows):

            upBound = self.upBound
            while len(solutions) < 100 and upBound <= len(rows):
                upBound = upBound + 1
                for i in range(0,last+1):
                    cols[i] = list(cols[i])
                    cols[i][3] = upBound
                    cols[i] = tuple(cols[i])
                matrix = DLX(cols, rows)
                solutions = self.solve(matrix)

            if len(solutions) < 100:
                self.typeoverlap_up = self.typeoverlap_up + 1
                for i in range(last+1,len(cols)):
                    cols[i] = list(cols[i])
                    cols[i][3] = self.typeoverlap_up
                    cols[i] = tuple(cols[i])
                matrix = DLX(cols, rows)
                solutions = self.solve(matrix)
            else:
                break

        print("Name: ", qdata["Full Name"], ", Product: ", product,", Rows: ", len(rows), ", Cols: ", len(cols), ", Solutions: ", end="")
        print(len(solutions))
        print("Unresolved: ", unresolved)

        bestSols = self.findBestSol(solutions, product, ailments)
        # try:
        #     bestSols = self.findBestSol(solutions, product, ailments)
        # except Exception as e:
        #     self.warn.displayWarningDialog("Error", f"error occured while finding best solution: {str(e)}")

        return bestSols, rows, cols, unresolved

    def solve(self, matrix):
        matrix.sols.connect(self.getDlxSols)
        self.cancel.connect(matrix.stop_)
        return matrix.dance()

    def findBestSol(self, solutions, product, ailments):
        # Get template folder path
        path = self.config.getDir("Formulation Sheets Directory") + "/"
        template_path = path + product + " Worksheet Template.xlsx"
        # Load the excel sheet
        workbook = load_workbook(filename=template_path)
        sheet = workbook.active
        # Get all required data to fill sheet
        ww_dict, assigned_vals = self.get_misc_items(sheet)

        target = self.config.getTarget(product)
        chosen = []
        solLen = len(solutions)
        _lenlst = [len(s) for s in solutions]
        maxlen, minlen = max(_lenlst), min(_lenlst)
        maxBenefits, leastBenefits = 0, 0

        j=0
        for solution in solutions:
            if self.stop:
                return None
            # send signal if the index of solution is a multiple of 100
            if j % 500 == 0:
                self.solsSorted(j, solLen)
            j = j+1

            vals = dd(list)
            benefits = 0
            blist = []
            for ingredient in solution:
                benefits_lst = [ingredient, []]
                # Finding information to calculate fit
                # Retrieve comodegenic rating
                _como = self.ingredients.loc[ingredient,self.comedogenicCol]

                if _como == "":
                    vals[ingredient].append(0)
                else:
                    vals[ingredient].append(self.comeConst.index(int(float(_como))))

                # Retrieve Viscocity
                key = self.ingredients.loc[ingredient,self.viscocityCol]
                try:
                    vals[ingredient].append(self.viscConst.index(key))
                except:
                    vals[ingredient].append(1)
                # Retrieve  absoption rate
                key = self.ingredients.loc[ingredient,self.absorptionCol]
                try:
                    vals[ingredient].append(self.absorbConst.index(key))
                except:
                    vals[ingredient].append(1)

                # Finding additional benefits
                for skinProb in self.ingredients.loc[ingredient,self.skinProbCol]:
                    if skinProb not in ailments:
                        benefits = benefits + 1
                        benefits_lst[1].append(skinProb)
                blist.append([benefits_lst[0], list(set(benefits_lst[1]))])
                if benefits > maxBenefits:
                    maxBenefits = benefits
                elif benefits < leastBenefits:
                    leastBenefits = benefits

            # Returns the percentage composition of each ingredient in the product
            composition = self.calc_ingredient_weight(solution, ww_dict, copy.deepcopy(assigned_vals))
            # Returns the point that this current solution occupies
            point = self.pointGen(composition, vals)
            # Returns the maximum distance from the target point and the distance to the point
            maxdist, dist = self.distFinder(target, point)
            # Calculate fit score (lower is better)
            fit = self.fitWeight * dist * 100 / maxdist
            # Calculate the score of the number of ingredients (lower is better)
            numIngred = self.numIngredWeight * (len(solution)-minlen) * 100 / (maxlen-minlen) if maxlen-minlen else 0
            # Calculating the score of the number of additional benefits (lower is better)
            addedBenefit = self.addedBenefitWeight * (maxBenefits - benefits) * 100 / (maxBenefits-leastBenefits) if maxBenefits-leastBenefits else 0
            score = fit + numIngred + addedBenefit
            # Need to find the additional benefits <----------------------------------------------------------------------

            if len(chosen) < self.maxSols:
                chosen.append((solution, score, blist))
            elif score < max([sol[1] for sol in chosen]):
                # remove the old maximum and add the solution to the list
                for i in range(len(chosen)):
                    if chosen[i][1] > score:
                        chosen[i] = (solution, score, blist)
                        break
        """
        print("Best solutions: ")
        for sol in chosen:
            print(sol)
        """
        return chosen

    def matrixGen(self, product, ailments, userCons):

        rows = []
        rownames = []
        # In the form [[ingredient types], [viscocity, Absorption rate, Comodegenic rating]]
        types = self.config.getProduct(product, "types")

        for index, ingredient in self.ingredients.iterrows():
            # Attain current stock, constraindications and ingredient type
            stock = ingredient[self.stockCol]
            ingredCons = ingredient[self.contrainsCol]
            type = ingredient[self.typeCol]

            # Filter the ingredients associated cures to contain keywords
            cures = ingredient[self.skinProbCol]

            # check if the ingredient is in stock, not a contraindication and useable
            if self.stockCheck(stock) \
              and self.contrainCheck(ingredCons, userCons) \
              and self.useablecheck(cures, ailments) \
              and self.typeCheck(types, type):

                # Create and append nodes for each row of the dlx matrix created
                nodes = self.dlxRowFormat(cures, ailments)
                rows.append((nodes, index))

        return rows, ailments

    def pointGen(self, composition, vals):

        point = []
        for i in range(3):
            val = 0
            for ingredient in vals.keys():
                val = val + vals[ingredient][i] * composition[ingredient]/100
            point.append(val)
        return point

    def distFinder(self,t, p):
        # t is the target point, p is the point
        # Find the distance of the point to the target point
        dist = math.sqrt((t[0]-p[0])**2 + (t[1]-p[1])**2 + (t[2]-p[2])**2)

        # find the maximum distance possible from the point
        maxX = max([5-t[0], t[0]]) # comedogenic rating
        maxY = max([len(self.config.getConst("viscosity"))-1-t[1], t[1]]) # viscocity
        maxZ = max([len(self.config.getConst("absorbency"))-1-t[2], t[2]]) # absorption
        maxdist = math.sqrt((t[0]-maxX)**2 + (t[1]-maxY)**2 + (t[2]-maxZ)**2)

        return maxdist, dist

    def stockCheck(self, stock):
        if not stock.lower() == "no":
                return True
        return False

    def contrainCheck(self, ingredCons, userCons):
        # ingredCons = list of ingredient contraindications
        # userCons = list of user contraindications
        if ingredCons and userCons:
            for ucon in userCons:
                for icon in ingredCons:
                    if ucon == icon:
                        return False
        return True

    def useablecheck(self, ingredSolves, problems):
        for i in ingredSolves:
            for k in problems:
                if k == i:
                    return True
        return False

    def typeCheck(self, types, type):
        for t1 in types:
            for t2 in type:
                if t1 == t2:
                    return True
        return False

    def dlxRowFormat(self, cures, problems):
        nodes = []
        for cure in cures:
            if cure in problems:
                nodes.append((problems.index(cure),None))
        return nodes

    def calc_ingredient_weight(self, solution, ww_dict, assigned_vals):
        """ Takes parameters from ingredient selector and calculates to ingredient weights
            inputs: solution - [ingredient names]
                    product type - product type string
                    self.ingredients - dataframe of ingredient database
        """
        # Fill main table entries
        for ingredient_name in solution:

            # Get ingredient types ???????????/// Waiting for answer to simplify column
            type_list = self.ingredients.loc[ingredient_name][self.typeCol]

            for ingredient_type in type_list:
                if "essential oil" in ingredient_type:
                    note = self.ingredients.loc[ingredient_name][self.EOnoteCol]
                    if note:
                        ingredient_type = "eo " + note
                    else:
                        ingredient_type = "eo middle"

                if ingredient_type in ww_dict.keys():
                    # Pop off and use first weight of ingredient type
                    if len(ww_dict[ingredient_type]) > 1:
                        assigned_vals[ingredient_name] = ww_dict[ingredient_type].pop(0)
                    # If only one is left then keep using that one
                    else:
                        assigned_vals[ingredient_name] = ww_dict[ingredient_type][0]

        # Scale to 100
        tot = sum(assigned_vals.values())
        if tot > 100:
            for key in assigned_vals.keys():
                assigned_vals[key] = round(assigned_vals[key] * 100/tot, 1)
        elif tot < 100:
            leftover = 100 - tot
            for key in assigned_vals.keys():
                assigned_vals[key] = round(assigned_vals[key] + leftover * assigned_vals[key]/tot, 1)
        else:
            pass
        return assigned_vals

    def get_misc_items(self, sheet):
        """ Creates the dicts and values needed for reallocation
            Returns: ww_dict - dict { ingredient type: w/w% }
                     assigned_dict - dict { ingredient type: w/w% } init. filled with all does not change names
        """
        ww_dict = dd(list)
        assigned_dict = {}

        i = self.SOF
        # Record the w/w% for all types
        while sheet[f"C{i}"].value != None:
            cell_ingredient = sheet[f"B{i}"].value.lower()
            cell_weight = sheet[f"D{i}"].value

            ww_dict[cell_ingredient].append(cell_weight)

            # Add fixed ingredient to assigned dict
            if ("doesn't change" in cell_ingredient) or ("does not change" in cell_ingredient):
                assigned_dict[cell_ingredient] = cell_weight

            i += 1
        return ww_dict, assigned_dict

    def solsSorted(self, i, max):
        state = "sorting"
        info = str(i) + " of " + str(max)
        self.stateChanged.emit(state, info, self.progress)

    @Slot(int)
    def getDlxSols(self, i):
        state = "finding"
        info = str(i) + " Solutions found"
        self.stateChanged.emit(state, info, self.progress)

    @Slot()
    def stop_(self):
        self.stop = True
        self.cancel.emit()
