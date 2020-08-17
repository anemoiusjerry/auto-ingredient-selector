import re
import os
import pandas as pd
from copy import copy
from PySide2.QtCore import QObject, Signal, Slot
# Spreadsheet library
from openpyxl import load_workbook
from datetime import *
from collections import defaultdict as dd

from config.configParser import FigMe
from BespokeAutoSystem.WarningRaiser import WarningRaiser

class FormulationFiller(QObject):
    SOF = 8
    launched = Signal(int)
    stateChanged = Signal(str, str, int)
    cancel = Signal()
    error = Signal(str)

    def __init__(self, ingredients_df, gdriveObject):
        QObject.__init__(self)
        self.stop = False

        self.ingredients_df = ingredients_df
        self.config = FigMe()
        self.gdriveObject = gdriveObject
        self.warn = WarningRaiser()

        self.errorStr = ""

    def process_all(self, results):
        for i, soln in enumerate(results):
            if self.stop:
                return
            name = soln["CustomerName"].title()
            product_type = soln["ProductType"].title()
            product_name = soln["ProductName"].lower()
            self.stateChanged.emit("writing", f"{product_name.title()} for {name}", i)

            isMale = False
            # Check for men's product
            if "courageous" in product_name or "men" in product_name or "man" in product_name:
                isMale = True

            try:
                self.write_to_template(soln["Ingredients"], name, product_type, isMale)
            except:
                self.errorStr += f"Failed to write {name}'s {product_type} formulation sheet. Check template has not changed"

        if self.errorStr != "":
            self.error.emit(self.errorStr)
            
        print("All form sheet generated.")

    def write_to_template(self, ingredients, customer_name, prod_type, isMale):
        """ Take information from ingredients selector and fill in
            formulation worksheet.
            inputs: ingredients - list of strings of ingredients
                    customer_name - str of customer name
                    prod_type - str of product type
        """
        # Get template path to read it
        path = self.config.getDir("Formulation Sheets Directory")
        formSheet_name = f"{prod_type} Worksheet Template"
        template_path = f"{path}/{formSheet_name}.xlsx"

        # Load the excel sheet
        try:
            if self.config.masterDict["gdrive"]:
                error_msg = "Failed to fetch formulation template from Google Drive."
                fh, file_id = self.gdriveObject.fetch_file(formSheet_name)
                workbook = load_workbook(fh)
            else:
                error_msg = "Failed to load formulation template."
                workbook = load_workbook(filename=template_path)
        except:
            self.errorStr += error_msg
        
        sheet = workbook.active

        ww_dict, phase_dict, realloc_dict, EOF = self.get_misc_items(sheet)
        assigned_vals = self.calc_ingredient_weight(ingredients, prod_type, self.ingredients_df)

        # Fill header info
        sheet["B1"] = customer_name
        sheet["B2"] = prod_type
        # Tag male product
        if isMale:
            sheet["A2"] = sheet["A2"].value + " [M]"
        if type(sheet["B5"].value) != int:
            sheet["B5"] = 100 # placeholder (varies)

        i=0
        # Keep trying to insert ingredients into slots if there still are ingredients
        while len(assigned_vals) > 0:
            # End loop if not enough slots
            if len(realloc_dict) <= 0 or i >= len(assigned_vals):
                sheet, EOF = self.too_many_slots(realloc_dict, sheet)
                realloc_dict = {}
                sheet = self.too_few_slots(assigned_vals, phase_dict, EOF, sheet)
                break

            ingredient_name = list(assigned_vals.keys())[i]
            # Insert ingredient into the correct row
            j=self.SOF
            while sheet[f"C{j}"].value != None and sheet[f"B{j}"].value != None:

                # Only fill w/w% if fixed ingredient
                if "fixed" in ingredient_name.lower():
                    if ingredient_name.lower() == sheet[f"B{j}"].value.lower():
                        # Fill w/w%
                        sheet[f"D{j}"] = assigned_vals[ingredient_name]

                        realloc_dict.pop(f"D{j}", None)
                        assigned_vals.pop(ingredient_name)
                        break

                else:
                    # Get ingredient types ???????????/// Waiting for answer to simply column
                    type_col_name = self.config.getColname("Ingredients Spreadsheet", "type")
                    type_list = self.ingredients_df.loc[ingredient_name][type_col_name]

                    for ingredient_type in type_list:

                        if "essential oil" in ingredient_type:
                            eo_note_col_name = self.config.getColname("Ingredients Spreadsheet", "EO note")
                            ingredient_type = "eo " + self.ingredients_df.loc[ingredient_name][eo_note_col_name].lower()

                        # Fill row in ingredient types match
                        if (ingredient_type == sheet[f"B{j}"].value.lower()):
                            # INCI
                            inci_col_name = self.config.getColname("Ingredients Spreadsheet", "inci")
                            sheet[f"A{j}"] = self.ingredients_df.loc[ingredient_name][inci_col_name]
                            # Ingredient Name
                            sheet[f"B{j}"] = ingredient_name
                            # w/w%
                            sheet[f"D{j}"] = assigned_vals[ingredient_name]
                            # Fill needs targted column if applicable
                            try:
                                if sheet[f"F{6}"].value.lower() == "needs targetting":
                                    prob_col_name = self.config.getColname("Ingredients Spreadsheet", "skin problem")
                                    sheet[f"F{6}"] = self.ingredients_df.loc[ingredient_name][prob_col_name]
                            except:
                                pass

                            realloc_dict.pop(f"D{j}", None)
                            assigned_vals.pop(ingredient_name)
                            break
                    # break out of inner while loop of ingredient is assigned
                    if ingredient_name not in assigned_vals.keys():
                        break
                j+=1

            # Move onto next if cannot write ingredient
            if ingredient_name in assigned_vals.keys():
                i+=1

        # Reallocate surplus %
        if len(realloc_dict) > 0:
            sheet, last_i = self.too_many_slots(realloc_dict, sheet)

        sheet = self.write_formulas(sheet)
        filename = customer_name + " - " + prod_type.title() + " Worksheet.xlsx"
        path = self.export_to_file(workbook, filename, customer_name)

    def calc_ingredient_weight(self, ingredients, prod_type, ingredients_df):
        """ Takes parameters from ingredient selector and calculates to ingredient weights
            inputs: ingredients - [ingredient names]
                    product type - product type string
                    ingredients_df - dataframe of ingredient database
        """
        # Get template folder path
        path = self.config.getDir("Formulation Sheets Directory") + "/"
        template_path = path + prod_type + " Worksheet Template.xlsx"

        # Load the excel sheet
        workbook = load_workbook(filename=template_path)
        sheet = workbook.active

        # Get all required data to fill sheet
        ww_dict, phase_dict, realloc_dict, EOF = self.get_misc_items(sheet)
        assigned_vals = {}

        # Fill main table entries
        for ingredient_name in ingredients:

            # Get ingredient types ???????????/// Waiting for answer to simplify column
            type_col_name = self.config.getColname("Ingredients Spreadsheet", "type")
            type_list = ingredients_df.loc[ingredient_name][type_col_name]

            for ingredient_type in type_list:
                # EO is labelled diff in form sheets
                if "essential oil" in ingredient_type:     
                    eo_note_col_name = self.config.getColname("Ingredients Spreadsheet", "EO note")                         # Haydens Fix
                    note = ingredients_df.loc[ingredient_name][eo_note_col_name]    #       |
                    if note:                                                        #       |
                        ingredient_type = "eo " + note                              #       |
                    else:                                                           #       \/
                        ingredient_type = "eo middle"                               # ---------------

                if ingredient_type in ww_dict:
                    # Pop off and use first weight of ingredient type
                    if len(ww_dict[ingredient_type]) > 1:
                        assigned_vals[ingredient_name] = ww_dict[ingredient_type].pop(0)
                    # If only one is left then keep using that one
                    else:
                        assigned_vals[ingredient_name] = ww_dict[ingredient_type][0]
        # Scale to 100
        assigned_vals = self.scale_water_weight(assigned_vals, ww_dict)
        return assigned_vals

    def scale_water_weight(self, assigned_vals, ww_dict):
        # Get fixed weight %
        fixed_weight = 0
        for ingredient, weight in ww_dict.items():
            if "fixed" in ingredient:
                fixed_weight += weight[0]

        # Scale to 100
        tot = sum(assigned_vals.values())
        target = 100 - fixed_weight
        if tot > target:
            for key in assigned_vals.keys():
                assigned_vals[key] = round(assigned_vals[key] * (target)/tot, 1)
        elif tot < target:
            leftover = target - tot
            for key in assigned_vals.keys():
                assigned_vals[key] = round(assigned_vals[key] + leftover * assigned_vals[key]/tot, 1)
        else:
            pass
        return assigned_vals


    def get_misc_items(self, sheet):
        """ Creates the dicts and values needed for reallocation
            Returns: ww_dict - dict { ingredient type: w/w% }
                     phase_dict - dict { ingredient type: phase }
                     realloc_dict - dict { cell ID: assigned % } Every time cell is filled, it is removed from realloc_dict
                     assigned_dict - dict { ingredient type: w/w% } init. filled with all does not change names
                     EOF - index of last row for ingredients
        """
        ww_dict = dd(list)
        phase_dict = {}
        realloc_dict = {}

        i = self.SOF
        # Record the w/w% for all types
        while sheet[f"C{i}"].value != None and sheet[f"B{i}"].value != None:
            cell_ingredient = sheet[f"B{i}"].value.lower()
            cell_weight = sheet[f"D{i}"].value
            # Guard against uninitialsed w/w
            if cell_weight == None:
                cell_weight = 0

            ww_dict[cell_ingredient].append(cell_weight)
            phase_dict[cell_ingredient] = sheet[F"C{i}"].value

            if not("fixed" in cell_ingredient):
                realloc_dict[f"D{i}"] = cell_weight

            i += 1
        EOF = i
        return ww_dict, phase_dict, realloc_dict, EOF

    def too_many_slots(self, realloc_dict, sheet):
        """ Reallocates extra percentages to other ingredients.
            Inputs: realloc_dict - { cell ID (D?): w/w% } ingredients still need to be allocated
            Returns: sheet with new values written and the new LAST row index.
        """
        # Write 0 to unallocated rows to mark for delete
        for cell in realloc_dict.keys():
            sheet[cell] = 0

        i = self.SOF
        while sheet[f"C{i}"].value != None:
            if sheet[f"D{i}"].value == 0:
                sheet.delete_rows(i)
            else:
                i+=1
        return sheet, i

    def too_few_slots(self, leftovers, phase_dict, EOF, sheet):
        """
            inputs: leftovers - dict { ingredient name: w/w% } of remaining ingredients
                    phase_dict - { ingredient type: phase }
                    EOF - last row of ingredients table
                    sheet - active sheet to write info to
        """

        for ingredient, weight in leftovers.items():
            # No need to add fixed ingredients as they are already in sheet
            if "fixed" in ingredient:
                continue

            # Get ingredient type and assign corresponding w/w%
            type_col_name = self.config.getColname("Ingredients Spreadsheet", "type")
            type_list = self.ingredients_df.loc[ingredient][type_col_name]
            for i_type in type_list:
                if i_type == "essential oil":
                    i_type = self.convert_eo_label(ingredient)

                if i_type in phase_dict.keys():
                    # Insert leftovers at end of sheet
                    sheet.insert_rows(EOF)
                    inci_col_name = self.config.getColname("Ingredients Spreadsheet", "inci")
                    sheet[f"A{EOF}"] = self.ingredients_df.loc[ingredient][inci_col_name]      # INCI name
                    sheet[f"B{EOF}"] = ingredient                                              # name
                    sheet[f"C{EOF}"] = phase_dict[i_type]                                      # Phase
                    sheet[f"D{EOF}"] = weight                                                  # w/w%
                    sheet[f"E{EOF}"] = f"=B5*D{EOF}/100"
                    # Copy over cell styles
                    for col_index in range(1,6):
                        sheet.cell(row=EOF, column=col_index).font = copy(sheet.cell(row=self.SOF, column=col_index).font)
                        sheet.cell(row=EOF, column=col_index).border = copy(sheet.cell(row=self.SOF, column=col_index).border)
                        sheet.cell(row=EOF, column=col_index).fill = copy(sheet.cell(row=self.SOF, column=col_index).fill)
                        sheet.cell(row=EOF, column=col_index).number_format = copy(sheet.cell(row=self.SOF, column=col_index).number_format)
                        sheet.cell(row=EOF, column=col_index).protection = copy(sheet.cell(row=self.SOF, column=col_index).protection)
                        sheet.cell(row=EOF, column=col_index).alignment = copy(sheet.cell(row=self.SOF, column=col_index).alignment)
                    EOF += 1
        return sheet

    def convert_eo_label(self, ingredient_name):  
        eo_note_col_name = self.config.getColname("Ingredients Spreadsheet", "EO note")
        note = self.ingredients_df.loc[ingredient_name][eo_note_col_name]    
        if note:                                                        
            ingredient_type = "eo " + note                              
        else:                                                           
            ingredient_type = "eo middle"  
        return ingredient_type

    def write_formulas(self, sheet):
        i=self.SOF
        while sheet[f"C{i}"].value != None:
            sheet[f"E{i}"] = f"=B5*D{i}/100"
            i+=1
        # Write totalising formula
        sheet[f"D{i}"] = f"=SUM(D{self.SOF}:D{i-1})"
        sheet[f"E{i}"] = f"=SUM(E{self.SOF}:E{i-1})"
        return sheet

    def export_to_file(self, workbook, filename, customer_name):
        """ Export as excel file to current working directory
        """
        save_path = self.config.getDir("Export Directory") + f"/{customer_name}/"
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        workbook.save(save_path + filename)
        print("Done exporting...")
        return save_path + filename

    @Slot()
    def stop_(self):
        self.stop = True
        self.cancel.emit()
